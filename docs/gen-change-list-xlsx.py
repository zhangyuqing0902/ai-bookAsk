# -*- coding: utf-8 -*-
"""生成《AI 问书 · 研发同步改动清单》xlsx（开会用，单页 + 页面截图）。

组织方式：**按功能模块**，不按端拆——「KP 下架」这一件事在机构后台和读者端分别是什么表现，
写在同一个功能点里，跨端的产品逻辑一次讲透。不同模块用浅色底区分。

正文内容在 docs/change-list-build/content.py（改文案只动那个文件）。
配图由 docs/change-list-build/shoot-changes.cjs 打线上截取对应交互态。
跑法：<venv>/bin/python docs/gen-change-list-xlsx.py
"""
import math
import os
import sys

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "AI问书-研发同步改动清单.xlsx")
SHOTS = os.path.join(HERE, "change-list-assets")
sys.path.insert(0, os.path.join(HERE, "change-list-build"))
from content import ITEMS  # noqa: E402

HEAD_FILL = PatternFill("solid", fgColor="4B57E8")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")
CTR_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
BOLD = Font(bold=True, size=10)
BODY = Font(size=10)
RED = "DC2626"

# 模块配色：(行底色, 分隔条底色, 分隔条字色)——都用浅色，只为分区，不抢正文
MOD_COLORS = {
    "知识产品 KP":   ("F4F6FF", "E3E7FF", "3942C9"),
    "会员与付费":     ("FFF8F0", "FFEAD6", "B45309"),
    "订单与退款":     ("F2FBF5", "DCF5E5", "15803D"),
    "数据看板":       ("F8F4FF", "EBE2FF", "6D28D9"),
    "机构与订阅":     ("FFFCF0", "FCF2CF", "A16207"),
    "权限与角色":     ("F1FAFF", "DBF0FD", "0369A1"),
    "导出与兑换码":   ("FEF4F9", "FBE2EF", "BE185D"),
    "登录与验证码":   ("EFFBF9", "D3F3EE", "0F766E"),
}
FALLBACK = ("FAFAFA", "EEEEEE", "444444")

DESC_COL_CHARS = 36   # 改动说明列一行大约放得下多少个汉字（按列宽 72 估）
IMG_H = 150           # 行内配图高度 px
MAX_SHOTS = 4


def ensure_tiers():
    """从原图派生压缩图（缺失或过期才重生成）。

    xlsx 嵌图保留的是原始文件字节、与显示尺寸无关：直接嵌 2x 原图会让文件翻好几倍。
    """
    import glob
    try:
        from PIL import Image
    except ImportError:
        print("⚠ 未安装 pillow，跳过压缩、直接用原图（xlsx 会偏大）")
        return
    d = os.path.join(SHOTS, "thumbs")
    os.makedirs(d, exist_ok=True)
    for src in sorted(glob.glob(os.path.join(SHOTS, "*.png"))):
        dst = os.path.join(d, os.path.basename(src))
        if os.path.exists(dst) and os.path.getmtime(dst) >= os.path.getmtime(src):
            continue
        im = Image.open(src).copy()
        im.thumbnail((1100, 1100), Image.LANCZOS)
        im.convert("P", palette=Image.ADAPTIVE, colors=256).save(dst, optimize=True)


def rich(text, size=10):
    """把 `**重点**` 标记渲染成红色加粗。

    Excel 单元格不认 Markdown，直接写 ** 会原样显示成星号；正文里用 ** 标重点很方便，
    所以在写入时统一转成真正的红色加粗片段。无 ** 时原样返回字符串。
    """
    if "**" not in text:
        return text
    blocks = []
    for i, seg in enumerate(text.split("**")):
        if not seg:
            continue
        f = InlineFont(b=True, color=RED, sz=size) if i % 2 else InlineFont(sz=size)
        blocks.append(TextBlock(f, seg))
    return CellRichText(*blocks)


def est_lines(text, per_line):
    """估算换行后占几行，用来定行高。** 标记不占宽度，先剥掉再算。"""
    plain = text.replace("**", "")
    return sum(max(1, math.ceil(len(seg) / per_line)) for seg in plain.split("\n"))


def put_shot(ws, row, col_letter, fname, px_h):
    p = os.path.join(SHOTS, "thumbs", fname)
    if not os.path.exists(p):
        p = os.path.join(SHOTS, fname)
    if not os.path.exists(p):
        print(f"   ⚠ 缺图 {fname}")
        return False
    img = XLImage(p)
    ratio = img.width / img.height
    img.height = px_h
    img.width = int(px_h * ratio)
    img.anchor = f"{col_letter}{row}"
    ws.add_image(img)
    return True


ensure_tiers()
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "改动说明"

HEADER = ["模块", "功能点", "改动说明", "影响端"] + [f"配图 {i + 1}" for i in range(MAX_SHOTS)]
WIDTHS = [14, 22, 72, 13] + [30] * MAX_SHOTS
ws.append(HEADER)
for c in range(1, len(HEADER) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill, cell.font, cell.border, cell.alignment = HEAD_FILL, HEAD_FONT, BORDER, CTR
for i, w in enumerate(WIDTHS):
    ws.column_dimensions[chr(65 + i)].width = w
ws.freeze_panes = "C2"          # 冻结表头 + 左两列，右滑看图时还知道是哪条

NCOL = len(HEADER)
r = 2
last_mod = None
missing = []
for mod, feat, desc, ends, shots in ITEMS:
    row_bg, bar_bg, bar_fg = MOD_COLORS.get(mod, FALLBACK)
    if mod != last_mod:                                   # 模块分隔条
        ws.cell(row=r, column=1, value=f"■ {mod}")
        for c in range(1, NCOL + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=bar_bg)
            cell.font = Font(bold=True, size=11, color=bar_fg)
            cell.border = BORDER
        ws.row_dimensions[r].height = 22
        r += 1
        last_mod = mod

    for c, v in enumerate([mod, feat, rich(desc), ends], start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border, cell.alignment, cell.font = BORDER, WRAP, BODY
    ws.cell(row=r, column=2).font = BOLD
    ws.cell(row=r, column=4).alignment = CTR_WRAP
    for c in range(1, NCOL + 1):                          # 整行铺模块底色
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=row_bg)
        ws.cell(row=r, column=c).border = BORDER

    for i, s in enumerate(shots[:MAX_SHOTS]):
        if not put_shot(ws, r, chr(69 + i), s, IMG_H):
            missing.append(s)
    ws.row_dimensions[r].height = max(est_lines(desc, DESC_COL_CHARS) * 14.5, IMG_H * 0.78)
    r += 1

wb.save(OUT)
mods = list(dict.fromkeys(m for m, *_ in ITEMS))
n_shots = sum(len(s[:MAX_SHOTS]) for *_, s in ITEMS)
print(f"✅ 已生成 {OUT}")
print(f"   {len(mods)} 个模块 / {len(ITEMS)} 个功能点 / {n_shots} 张配图 / 单 sheet")
if missing:
    print(f"   ⚠ 缺失配图 {len(missing)} 张：{', '.join(sorted(set(missing)))}")
