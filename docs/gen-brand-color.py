# -*- coding: utf-8 -*-
"""生成「机构品牌色影响范围」详版清单 xlsx（用 venv python 跑：docs/feature-list-build/.venv/bin/python）。
口径：从 UX / 产品 / UI 设计角度，穷尽机构前台 H5 + 机构后台 PC 哪些随机构品牌色变、哪些不变；
     平台超管 PC 作为作用域边界单列一条（整端固定平台色，不参与机构换肤）。
现状：平台超管·机构详情·品牌外观 可配主色/辅色，但前台/后台均无代码消费，本表是规格 + 技术落地清单。

━━━━━━━━━━━━ 接入前风险提示（拍板换肤方案前必读）━━━━━━━━━━━━
1) 「换个变量就能改色」是假象。大量位置连 var() 都不是、直接写死 hex：
   .btn-primary / .btn-amber 主按钮渐变(styles.css:179-181)、.orb 知识核 5 层渐变、纸书封面渐变、
   留存率渐变文字(admin-app.css:2384)、全部折线图主色(#4B57E8 共 33 处)、导出 xlsx 表头色(exportCsv.ts:120)；
   另有 rgba(75,87,232,…) 半透品牌靛 45 处散落各文件。
   建议先做一轮硬编码收敛（统一改 color-mix(in srgb, var(--org-primary) X%, transparent)）再谈换色，
   否则「换肤」只换到一半，页面会出现新品牌色与残留电光靛同屏。
2) 白字压主色处存在对比度风险：主按钮、日历选中日、TTS 播放键、答案反馈用户气泡、导出 xlsx 表头，
   在浅色品牌色下前景白字对比度会跌破 3:1。取色处需加亮度校验（L>0.6 时前景自动切 --ink）或限制可选色域。
   其中 Excel 表头没有运行时降级余地（导出件离线打开），最危险，必须在取色环节拦住。
3) --indigo-soft 被 57 处当选中底/hover 底，且常与 --indigo-ink 配成「淡底 + 深字」。
   深色品牌色下必须用 color-mix 按固定比例生成 soft / base / ink 三阶，不能让机构自己手填三个色值。
4) 品牌色若落在 --jade(#15B080) 或 --terra(#E5533B) 附近，会与「已发布=绿 / 退款=红」等语义色同屏冲突，
   取色控件应拒绝红/绿色相区间或至少给出警告。
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

H = ['序号', '端', '页面 / 模块', '位置 / 元素', '色用法', '是否随品牌色变', '当前实现', '建议变量 / 说明']

# (端, 页面, 位置, 色用法, 是否随品牌变, 当前实现, 建议/说明)
ROWS = [
    # —— 机构前台 H5 · 需随品牌色变 ——
    ['机构前台 H5', '登录落地页', '手机号登录按钮', '主色填充', '✅ 需变', '硬编码 indigo', '--org-primary'],
    ['机构前台 H5', '登录落地页', '知识核动效球', '主→辅渐变', '✅ 需变', '硬编码', '--org-grad'],
    ['机构前台 H5', '登录落地页', '「问书」渐变标题', '主→辅渐变文字', '✅ 需变', '--grad', '--org-grad'],
    ['机构前台 H5', '手机验证码登录', '获取验证码 / 登录按钮', '主色', '✅ 需变', '--indigo', '--org-primary'],
    ['机构前台 H5', 'AI 会话', '发送按钮', '主色填充', '✅ 需变', '--indigo', '--org-primary'],
    ['机构前台 H5', 'AI 会话', '能力开关选中态(深度思考/智能搜索)', '主色软底 + 主色字', '✅ 需变', 'indigo 软底', '--org-primary / -soft'],
    ['机构前台 H5', 'AI 会话', '语音输入音浪动画', '主色', '✅ 需变', '硬编码', '--org-primary'],
    ['机构前台 H5', 'AI 会话', '视频/音频播放进度条', '主色填充', '✅ 需变', '--indigo', '--org-primary'],
    ['机构前台 H5', 'AI 会话', '推荐追问 / 可点链接', '主色', '✅ 需变', '--indigo', '--org-primary'],
    ['机构前台 H5', 'AI 会话', '音频封面渐变底', '主→辅渐变', '✅ 需变', '硬编码', '--org-grad'],
    ['机构前台 H5', 'AI 会话', '欢迎态「全库检索」说明 chip', '主色淡底 + 主色字', '✅ 需变', 'indigo-soft', '--org-primary / -soft'],
    ['机构前台 H5', '多模态付费墙', '永享买断按钮', '价值色(珊瑚)', '⚠️ 建议固定', '.btn btn-amber 硬编码渐变 #FF9879→#F2563B', '价值色体系，与会员价值色一致；若随品牌色会出现「品牌色按钮 + 珊瑚角标」双主色打架；Chat.tsx:568/572'],
    ['机构前台 H5', '会员中心', '会员卡光球', '主色', '✅ 需变', '硬编码', '--org-primary'],
    ['机构前台 H5', '会员中心', '开通会员主按钮', '主色', '✅ 需变', '--indigo', '--org-primary'],
    ['机构前台 H5', '微信支付收银台', '顶部 logo 动画球', '主→辅渐变', '✅ 需变', '硬编码', '--org-grad'],
    ['机构前台 H5', '支付成功页', '成功对勾同心圆光环', '主色', '✅ 需变', '硬编码', '--org-primary'],
    ['机构前台 H5', '我的纸书', 'KP 封面渐变底', '主→辅渐变', '✅ 需变', '硬编码', '--org-grad'],
    ['机构前台 H5', '我的 / 个人中心', '用户卡 / 会员标强调', '主色', '✅ 需变', '硬编码', '--org-primary'],
    ['机构前台 H5', '兑换码', '立即兑换按钮', '主色', '✅ 需变', '--indigo', '--org-primary'],
    ['机构前台 H5', '全局', '主按钮 / 选中态 / 单选复选勾选', '主色', '✅ 需变', '硬编码渐变 #7179F8→#3C49D6', '--org-primary；.btn-primary 从不引用 --indigo，须先收敛硬编码；styles.css:179-180'],
    # —— 机构前台 H5 · 0712~0718 六批次新增页面/组件 ——
    ['机构前台 H5', 'KP 判活入口页(新增)', '下架/失效提示图标底+图标', '语义琥珀/珊瑚', '❌ 不需变', 'amber-soft / terra-soft', '拦截页语义色固定；screens/KpGate.tsx:72-81'],
    ['机构前台 H5', '我的纸书', '整卡降透明(已下架/已失效)', '中性降透明', '❌ 不需变', '.bk-dim opacity .55', '不着色，只降透明；mobile-app.css:1363'],
    ['机构前台 H5', '我的纸书', '「已解锁」封面角标', '语义成功绿', '❌ 不需变', '.bk-tag-unlock var(--jade)', 'mobile-app.css:1367'],
    ['机构前台 H5', '我的纸书', '「已下架/已失效」中性胶囊', '中性灰', '❌ 不需变', '.bk-st rgba(23,26,33,.055)', 'mobile-app.css:1369，0718 已去彩色'],
    ['机构前台 H5', '我的永享', '「已失效」珊瑚标 / 「已下架」琥珀标', '语义色', '❌ 不需变', '.yx-dead / .yx-off', 'mobile-app.css:1377/1379'],
    ['机构前台 H5', '我的永享', '失效封面置灰', '灰度滤镜', '❌ 不需变', '.yx-cover.dead grayscale(1)', 'mobile-app.css:1376'],
    ['机构前台 H5', '开通会员(双模式)', '权益速览 chip', '主色字 + 主色淡描边', '✅ 需变', 'var(--indigo-ink) + rgba(75,87,232,.14)', '--org-primary；mobile-app.css:1384'],
    ['机构前台 H5', '开通会员(双模式)', '套餐卡选中态(描边+底+对勾)', '价值橙', '⚠️ 建议固定', '.mb-plan.sel var(--amber)', '会员体系统一橙；mobile-app.css:1389/1393'],
    ['机构前台 H5', '开通会员(双模式)', '「首月特惠」角标', '珊瑚渐变', '⚠️ 建议固定', '.mb-plan-badge 硬编码 #FF9879→#F2563B', 'mobile-app.css:1390'],
    ['机构前台 H5', '我的订单 / 订单详情', '退款处理中状态字', '主色', '✅ 需变', '.refund-state.processing var(--indigo)', '--org-primary；mobile-app.css:1343'],
    ['机构前台 H5', '微信扫码登录', '扫码状态圆形图标', '主色淡底 + 主色', '✅ 需变', '.wxscan-mask-ic', '--org-primary/-soft；mobile-app.css:1303-1308'],
    ['机构前台 H5', '微信扫码登录', '「返回」文字链', '主色深阶', '✅ 需变', '.wxscan-back var(--indigo-ink)', '--org-primary；mobile-app.css:1330-1332'],
    ['机构前台 H5', 'AI 会话', '答案反馈标签选中态', '主色淡底 + 主色字', '✅ 需变', '.fbk-tag.on', '--org-primary/-soft；mobile-app.css:672-675'],
    # —— 机构后台 PC · 需随品牌色变 ——
    ['机构后台 PC', '登录页', '登录按钮', '主色', '✅ 需变', '--indigo', '--org-primary'],
    ['机构后台 PC', '登录页', '左侧品牌区 blob / 知识核', '主→辅渐变', '✅ 需变', '硬编码', '--org-grad'],
    ['机构后台 PC', '全局·侧栏', '菜单激活态(左竖条+底色+文字)', '主色', '✅ 需变', '--side-active / --indigo', '--org-primary'],
    ['机构后台 PC', '全局', '主操作按钮(新建/保存/导出)', '主色填充', '✅ 需变', '硬编码渐变 #7179F8→#3C49D6', '--org-primary；.btn-primary 从不引用 --indigo，hover 另有硬编码 #413bab；styles.css:179-180'],
    ['机构后台 PC', '全局', 'Tab 选中下划线', '主色', '✅ 需变', '--indigo', '--org-primary'],
    ['机构后台 PC', '全局', '链接 / 可点文字', '主色', '✅ 需变', '--indigo', '--org-primary'],
    ['机构后台 PC', '全局', '开关 / 单选 / 复选 选中态', '主色', '✅ 需变', '--indigo', '--org-primary'],
    ['机构后台 PC', '全局', '输入框聚焦边框', '主色', '✅ 需变', '--indigo', '--org-primary'],
    ['机构后台 PC', '全局', '上传 / 反馈处理进度条填充', '主色', '✅ 需变', '--indigo', '--org-primary（.up-item-bar i / .fbar i）'],
    ['机构后台 PC', '订阅与配额', '配额用量进度条填充', '语义三色(绿/橙/红)', '❌ 不需变', '.quota-fill ok/warn/bad = jade/amber/terra', '用量健康度语义，不能被品牌色吃掉；admin-app.css:1815-1817'],
    ['机构后台 PC', '主控台', '趋势图 / KPI 图表主色系', '主色系', '✅ 需变', 'Recharts 硬编码', '--org-primary'],
    ['机构后台 PC', '主控台', '当前订阅卡强调 / 徽标', '主色', '✅ 需变', '硬编码', '--org-primary'],
    ['机构后台 PC', '数据看板', '留存/来源/分布等图表主色系', '主色系', '✅ 需变', 'Recharts 硬编码', '--org-primary'],
    ['机构后台 PC', '数据看板', '提问关键词云·色深', '主色系', '✅ 需变', '硬编码', '--org-primary'],
    # —— 机构后台 PC · 0712~0718 六批次新增页面/组件 ——
    ['机构后台 PC', '数据看板', 'KPI 卡 hover 顶条 + 右上角光晕', '主渐变 / 主色半透', '✅ 需变', 'var(--grad) + rgba(75,87,232,.09)', '--org-grad；admin-app.css:2301-2320'],
    ['机构后台 PC', '数据看板', '段标题左侧渐变竖条', '主渐变', '✅ 需变', 'var(--grad)', '--org-grad；admin-app.css:2360-2368'],
    ['机构后台 PC', '数据看板', '主题 Tab 选中字 + 渐变指示条 + hover 底', '主色深阶 / 主渐变', '✅ 需变', 'var(--indigo-ink) / var(--grad)', '--org-primary/-grad；admin-app.css:2374-2376'],
    ['机构后台 PC', '数据看板·留存三卡', '留存率大数字(渐变文字)', '主→辅渐变文字', '✅ 需变', '硬编码 #3d6ff5→#8b6cf6', '--org-grad；admin-app.css:2380-2387'],
    ['机构后台 PC', '数据看板·留存三卡', '留存比例条 轨道 + 填充', '主色淡底 + 主渐变', '✅ 需变', 'var(--indigo-soft) + 硬编码渐变', '--org-primary-soft/--org-grad；admin-app.css:2389-2399'],
    ['机构后台 PC', '数据看板·留存三卡', '「待成熟」占位卡', '中性虚线灰', '❌ 不需变', '.ret-pending 中性', '占位态不着色；admin-app.css:2401-2418'],
    ['机构后台 PC', '数据看板·来源分布', '环形图弧段 + 投影', '主→辅渐变', '✅ 需变', '硬编码 #3D6FF5→#8B6CF6', '--org-grad；DataBoard.tsx:294-301'],
    ['机构后台 PC', '数据看板·来源分布', '图例色块 + 行 hover 底', '主→辅渐变 / 主色淡底', '✅ 需变', '硬编码渐变 / var(--indigo-soft)', '--org-grad/-soft；DataBoard.tsx:310'],
    ['机构后台 PC', '数据看板', '新增用户/提问量迷你趋势折线', '主色系', '✅ 需变', '硬编码 #4B57E8', '--org-primary；DataBoard.tsx:283/369'],
    ['机构后台 PC', '全局·图表', '折线图 hover 竖向指引线', '主色', '✅ 需变', '硬编码 #4B57E8', '--org-primary；ui-admin/LineChart.tsx:55（三端共用）'],
    ['机构后台 PC', '数据看板', '「按注册时间」标签', '主色淡底 + 主色字', '✅ 需变', '.tag-indigo', '--org-primary/-soft；styles.css:196'],
    ['机构后台 PC', '数据看板', '区间分析条 / 排行行 hover', '主色淡底', '✅ 需变', 'var(--indigo-soft)', '--org-primary-soft；admin-app.css:2152/2194'],
    ['机构后台 PC', '订阅与配额', '订阅方案行选中态(描边+底+对勾)', '主色 + 主色淡底', '✅ 需变', '.sub-plan-row.on / .spr-check', '--org-primary/-soft；admin-app.css:552-586'],
    ['机构后台 PC', '订阅与配额', '额度卡 hover/focus 态 + 图标圆底', '主色描边 + 主色淡底', '✅ 需变', '.quota-card / .qc-ic', '--org-primary/-soft；admin-app.css:596-600'],
    ['机构后台 PC', '订阅与配额', '额度步进器 +/− hover、输入聚焦环', '主色描边 + 主色半透光环', '✅ 需变', '.qty-stepper var(--indigo) + rgba(75,87,232,.14)', '--org-primary；admin-app.css:608/612'],
    ['机构后台 PC', '订阅与配额', '说明条圆点 / 规则列表项目符', '主色', '✅ 需变', '.quota-note::before / .sub-tip-rules li::before', '--org-primary；admin-app.css:618/671'],
    ['机构后台 PC', '订阅与配额', '「占用量」额度类型标', '主色淡底 + 主色字', '✅ 需变', '.quota-kind.occupancy', '--org-primary/-soft；admin-app.css:2263（消耗量用 amber 不变）'],
    ['机构后台 PC', '加油包抽屉', '空态插画', '主色描边+填充', '✅ 需变', '硬编码 #4B57E8 多处', '--org-primary；ui-admin/SubPackDrawer.tsx:49-55'],
    ['机构后台 PC', 'Agent 列表 / 详情', '「机构」类型标', '主色淡底 + 主色字', '✅ 需变', '.tag-indigo（0718 从绿改靛）', '--org-primary/-soft；AgentList.tsx:9'],
    ['机构后台 PC', 'Agent 列表', 'Agent 头像渐变底', '主→辅渐变', '✅ 需变', '硬编码 #7c8bf5→#5562d8', '--org-grad；AgentList.tsx:9'],
    ['机构后台 PC', 'C 端用户详情', '用户头像光球', '主→辅渐变', '✅ 需变', '硬编码 #7c8bf5→#4b57e8', '--org-grad；CUserDetail.tsx:69'],
    ['机构后台 PC', '知识 KP 详情', '内容 Tab 选中下划线', '主色', '✅ 需变', '.kpd-tab.on::after var(--indigo)', '--org-primary；proto-admin.css:247'],
    ['机构后台 PC', '知识 KP 列表 / 详情', '发布状态标(已发布/已下架/草稿)', '语义绿/琥珀/灰描边', '❌ 不需变', 'tag-jade / tag-amber / tag-line', '0718 已把「已发布」从电光靛改语义绿，勿再回收品牌色；proto-admin.css:506-511'],
    ['机构后台 PC', '全局·表单', '禁用置灰统一标准', '中性纸白底 + 中性边', '❌ 不需变', '.inp2.disabled / .sel-disabled', '禁用态必须中性，品牌色会被误读成可用；proto-admin.css:436-440'],
    ['机构后台 PC', '全局·下拉', '选项选中态 / 禁用项', '主色淡底+主色字 / 中性', '✅ 需变 · 禁用项不变', '.dd-opt.on / .dd-opt.disabled', '--org-primary/-soft；admin-app.css:68-70/989'],
    ['机构后台 PC', '系统配置', '前台访问地址输入框聚焦态', '主色描边 + 主色半透光环', '✅ 需变', '.domain-input:focus-within', '--org-primary；admin-app.css:2251'],
    ['机构后台 PC', '全局·上传', '拖拽区 hover/激活态', '主色描边+淡底+字', '✅ 需变', '.up-drop', '--org-primary/-soft；admin-app.css:1892-1896'],
    ['机构后台 PC', '答案反馈工作台', '用户气泡渐变底', '主渐变', '✅ 需变', 'var(--grad-bubble)', '--org-grad；proto-admin.css:148'],
    ['机构后台 PC', '答案反馈工作台', '处理中状态 + 进度条', '主色', '✅ 需变', '.fstat.ing / .fbar i', '--org-primary；proto-admin.css:255/262'],
    ['机构后台 PC', '账号与权限', '角色列表选中态', '主色淡底 + 主色字', '✅ 需变', '.role.on', '--org-primary/-soft；proto-admin.css:308/405'],
    ['机构后台 PC', '账号与权限', '权限嵌套子项连接线', '中性灰虚线', '❌ 不需变', '.perm-child-wrap::before dashed var(--ink-3)', '已刻意去掉子项紫色底、只留缩进+虚线；proto-admin.css:423-427'],
    ['机构后台 PC', '账号与权限', '读/写权限档位块', '主色', '✅ 需变', '.lvl-seg b.on.write', '--org-primary；proto-admin.css:88/100'],
    ['机构后台 PC', '全局·时间筛选', '分段控件选中 / 日历选中日 / 区间内', '主色 + 主色淡底', '✅ 需变', '.seg b.on / .cal-d.sel / .cal-d.inrange', '--org-primary/-soft；proto-admin.css:208/451-467'],
    ['机构后台 PC', '全局·时间筛选', '自定义区间回显 chip', '主色淡底 + 主色字', '✅ 需变', '.dr-applied', '--org-primary/-soft；admin-app.css:996-1001'],
    ['机构后台 PC', '全局·表格', '排序箭头激活态', '主色', '✅ 需变', '.tbl th.sortable.asc/.desc .sort i', '--org-primary；proto-admin.css:392-393'],
    ['机构后台 PC', '全局·顶栏', '用户菜单 hover / 头像 / 空态插画', '主色淡底 + 主色深阶', '✅ 需变', '.menu-row:hover / .pc-avatar / .eill', '--org-primary/-soft；admin-app.css:174-281'],
    ['机构后台 PC', '个人中心', '头像裁剪缩放滑块 / 密码强度提示', '主色', '✅ 需变', '.acrop-zoom accent-color / .cred-pwd', '--org-primary；admin-app.css:429/708'],
    ['机构后台 PC', '知识 KP 详情·媒体', '播放键/进度条/倍速/波形激活', '主色', '✅ 需变', '.mv-pp / .mv-track i / .mv-bars span.on', '--org-primary；admin-app.css:1417-1532'],
    ['机构后台 PC', '导出(两后台通用)', '导出 xlsx 表头填充 / 标题栏底', '主色深阶实底 + 白字', '✅ 需变', '硬编码 argb FF3730A3 / FFF2F4FF', '--org-primary；ui-admin/exportCsv.ts:87/120——导出件是带出机构的交付物，最该跟品牌走'],
    # —— 平台超管 PC · 作用域边界声明 ——
    ['平台超管 PC', '全局', '所有主色 / 渐变元素', '平台自身品牌', '❌ 不需变', '固定 --indigo / --grad', '超管是平台侧界面、不属于任何机构，必须固定平台色；19 个视图与机构后台共用 proto-admin.css + admin-app.css，换变量时勿误伤'],
    # —— 不随品牌色变（保持固定，重要！）——
    ['三端通用', '全局', '成功 / 已解锁 / 上架 标识', '语义成功绿', '❌ 不需变', '--jade', '语义色固定——变了会失去「成功」直觉'],
    ['三端通用', '全局', '删除 / 危险 / 错误 / 退款', '语义错误红', '❌ 不需变', '--terra', '语义色固定——警示性不能被品牌色稀释'],
    ['机构前台 H5', '登录 / 支付', '微信登录 · 微信支付 按钮', '微信品牌绿', '❌ 不需变', '微信绿', '微信官方品牌色，识别度与合规要求，不可改'],
    ['三端通用', '全局', '正文 / 标题 文字', '中性墨色', '❌ 不需变', '--ink / --ink-2', '中性色，保证可读性，品牌色只做点缀不做正文'],
    ['三端通用', '全局', '页面 / 卡片 背景', '奶白 / 白', '❌ 不需变', '--surface / --paper', '中性底；品牌色仅用于强调，不做大面积背景'],
    ['三端通用', '全局', '边框 / 分隔线', '中性灰', '❌ 不需变', '--line / --line-2', '中性结构色，跨机构统一'],
    ['机构前台 H5', 'AI 会话', '资源类型角标(图 / 音 / 视)', '类型区分色', '❌ 不需变', '固定', '类型标识跨机构一致，便于用户快速识别'],
    ['三端通用', '会员体系', '会员标识 / 会员价值色', '价值橙', '⚠️ 建议固定', '--amber', '会员是平台级权益体系，建议跨机构统一色（可商议）'],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '品牌色影响清单'
head_fill = PatternFill('solid', fgColor='4B57E8')
head_font = Font(bold=True, color='FFFFFF', size=11)
change_fill = PatternFill('solid', fgColor='EAF7F0')   # 需变：淡绿
keep_fill = PatternFill('solid', fgColor='F4F4F6')     # 不需变：淡灰
thin = Side(style='thin', color='DDDDDD')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical='center')
center = Alignment(horizontal='center', vertical='center')

ws.append(H)
for c in range(1, len(H) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = head_fill; cell.font = head_font; cell.alignment = center; cell.border = border

for i, r in enumerate(ROWS, start=2):
    ws.cell(row=i, column=1, value=i - 1)
    for c, v in enumerate(r, start=2):
        ws.cell(row=i, column=c, value=v)
    fill = change_fill if r[4].startswith('✅') else keep_fill
    for c in range(1, len(H) + 1):
        cell = ws.cell(row=i, column=c)
        cell.border = border
        cell.alignment = center if c in (1, 6) else wrap
        if c == 6:
            cell.fill = fill

widths = [6, 13, 22, 30, 20, 14, 22, 30]
for col, w in zip('ABCDEFGH', widths):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'brand-color-impact.xlsx')
wb.save(out)
n_change = sum(1 for r in ROWS if r[4].startswith('✅'))
n_fix = sum(1 for r in ROWS if r[4].startswith('❌'))
n_advise = sum(1 for r in ROWS if r[4].startswith('⚠️'))
print(f'生成完成: {out}')
print(f'共 {len(ROWS)} 条：✅ 需变 {n_change} · ❌ 不需变 {n_fix} · ⚠️ 建议固定 {n_advise}')
from collections import Counter
for side, cnt in Counter(r[0] for r in ROWS).items():
    print(f'  {side}: {cnt} 条')
